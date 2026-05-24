"""Import Microsoft Forms Excel responses into response_sessions + responses.

Usage (from backend/):
  pip install openpyxl
  python import_responses_excel.py --file data/responses.xlsx --dry-run
  python import_responses_excel.py --file data/responses.xlsx --commit

Place your export at backend/data/responses.xlsx (or pass --file).
Does not delete existing responses. Duplicate respondent names are allowed; use --on-duplicate suffix to uniquify labels when importing.
"""

from __future__ import annotations

import argparse
import asyncio
import difflib
import re
import sys
from collections import Counter
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from answer_normalize import AnswerValidationError, cell_to_raw, normalize_answer
from database import get_session_factory
from models import FormPage, Question, Response, ResponseSession
from soft_delete import active_form_page, active_question
from respondent_name import find_name_question

DEFAULT_PATH = Path(__file__).resolve().parent / "data" / "responses.xlsx"

META_HEADERS = frozenset(
    {
        "id",
        "response id",
        "start time",
        "completion time",
        "last submitted time",
        "last modified time",
        "email",
        "name",
        "language",
        "responder",
        "responder name",
        "responder email",
        "user name",
        "username",
    }
)

def normalize_header(text: str) -> str:
    s = str(text or "").replace("\xa0", " ").replace("\u200b", "")
    s = re.sub(r"\s+", " ", s.strip().lower())
    s = re.sub(r"^\[section\]\s*", "", s, flags=re.IGNORECASE)
    return s


def is_meta_column(header: str) -> bool:
    return normalize_header(header) in META_HEADERS


def build_question_lookup(questions: list[Question]) -> dict[str, Question]:
    lookup: dict[str, Question] = {}
    for q in questions:
        key = normalize_header(q.question_text)
        if key and key not in lookup:
            lookup[key] = q
    return lookup


def match_column_to_question(header: str, lookup: dict[str, Question]) -> Question | None:
    key = normalize_header(header)
    if not key or is_meta_column(header):
        return None
    if key in lookup:
        return lookup[key]
    # Microsoft Forms truncates long headers in Excel; match by shared prefix.
    best: Question | None = None
    best_len = 0
    for qkey, question in lookup.items():
        prefix_len = 0
        for a, b in zip(key, qkey):
            if a != b:
                break
            prefix_len += 1
        if prefix_len >= 40 and prefix_len > best_len:
            best = question
            best_len = prefix_len
        elif prefix_len >= min(25, len(key), len(qkey)) and prefix_len > best_len:
            best = question
            best_len = prefix_len
    if best is not None:
        return best
    close = difflib.get_close_matches(key, lookup.keys(), n=1, cutoff=0.88)
    if close:
        return lookup[close[0]]
    return None


def pick_respondent_name(
    row_meta: dict[str, str | None],
    answers_by_qid: dict[int, str],
    name_question_id: int | None,
    row_index: int,
    used_names: set[str],
    on_duplicate: str,
) -> str | None:
    candidates: list[str] = []
    if name_question_id is not None and name_question_id in answers_by_qid:
        candidates.append(answers_by_qid[name_question_id])
    for key in ("name", "email", "responder", "responder name", "username"):
        val = row_meta.get(key)
        if val:
            candidates.append(val)
    ms_id = row_meta.get("id")
    if ms_id:
        candidates.append(f"import-{ms_id}")
        candidates.append(f"respondent-{ms_id}")
    candidates.append(f"import-row-{row_index}")

    for raw in candidates:
        name = re.sub(r"\s+", " ", raw.strip())[:200]
        if len(name) < 2:
            continue
        if name not in used_names:
            used_names.add(name)
            return name
        if on_duplicate == "skip":
            return None
        if on_duplicate == "suffix":
            n = 2
            while True:
                candidate = f"{name} ({n})"[:200]
                if candidate not in used_names:
                    used_names.add(candidate)
                    return candidate
                n += 1
    return None


def find_name_question_id(questions: list[Question]) -> int | None:
    q = find_name_question(questions)
    return q.id if q is not None else None


async def run_import(
    file_path: Path,
    *,
    sheet: str | None,
    commit: bool,
    on_duplicate: str,
    strict: bool,
    limit: int | None,
) -> int:
    if not file_path.is_file():
        print(f"File not found: {file_path}", file=sys.stderr)
        return 1

    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    if ws.max_row < 2:
        print("Excel file has no data rows.", file=sys.stderr)
        wb.close()
        return 1

    headers = [
        str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else ""
        for c in range(1, ws.max_column + 1)
    ]
    col_meta: dict[int, str] = {}
    col_question: dict[int, Question] = {}

    async with get_session_factory()() as session:
        result = await session.execute(
            select(Question)
            .join(FormPage, Question.page_id == FormPage.id)
            .where(active_form_page(), active_question())
            .options(selectinload(Question.page))
            .order_by(FormPage.order_index, Question.order_index)
        )
        questions = list(result.scalars().all())
        if not questions:
            print("No questions in database. Run import_microsoft_form.py first.", file=sys.stderr)
            return 1

        lookup = build_question_lookup(questions)
        name_question_id = find_name_question_id(questions)

        unmapped: list[str] = []
        for idx, header in enumerate(headers):
            if not header:
                continue
            if is_meta_column(header):
                col_meta[idx] = normalize_header(header)
                continue
            q = match_column_to_question(header, lookup)
            if q is None:
                unmapped.append(header)
            else:
                col_question[idx] = q

        existing = await session.execute(select(ResponseSession.respondent_name))
        used_names = {n for (n,) in existing.all() if n}

        stats = Counter()
        pending_sessions: list[tuple[ResponseSession, list[Response]]] = []

        for row_num in range(2, ws.max_row + 1):
            if limit is not None and stats["rows_seen"] >= limit:
                break
            stats["rows_seen"] += 1

            row = [ws.cell(row_num, c).value for c in range(1, len(headers) + 1)]
            if all(cell is None or str(cell).strip() == "" for cell in row):
                stats["empty_rows"] += 1
                continue

            row_meta: dict[str, str | None] = {}
            answers_by_qid: dict[int, str] = {}

            for col_idx, cell in enumerate(row):
                raw = cell_to_raw(cell)
                if raw is None:
                    continue
                if col_idx in col_meta:
                    row_meta[col_meta[col_idx]] = raw
                    continue
                question = col_question.get(col_idx)
                if question is None:
                    continue
                try:
                    answers_by_qid[question.id] = normalize_answer(
                        question, raw, strict=strict
                    )
                except AnswerValidationError as exc:
                    stats["invalid_cells"] += 1
                    if strict:
                        print(f"Row {row_num}, column {headers[col_idx]!r}: {exc}", file=sys.stderr)
                        return 1
                    stats.setdefault("warnings", 0)
                    stats["warnings"] += 1

            if not answers_by_qid:
                stats["no_answers"] += 1
                continue

            respondent_name = pick_respondent_name(
                row_meta,
                answers_by_qid,
                name_question_id,
                row_num,
                used_names,
                on_duplicate,
            )
            if respondent_name is None:
                stats["skipped_duplicate"] += 1
                continue

            session_id = uuid4()
            rs = ResponseSession(
                session_id=session_id,
                respondent_name=respondent_name,
                submitted_by_user_id=None,
            )
            response_rows = [
                Response(
                    session_id=session_id,
                    question_id=qid,
                    answer_value=val,
                )
                for qid, val in answers_by_qid.items()
            ]
            pending_sessions.append((rs, response_rows))
            stats["imported"] += 1

        print(f"Sheet: {ws.title!r}")
        print(f"Questions in DB: {len(questions)}")
        print(f"Mapped columns: {len(col_question)}")
        if unmapped:
            print(f"Unmapped columns ({len(unmapped)}) — ignored:")
            for h in unmapped[:15]:
                print(f"  - {h[:100]}{'…' if len(h) > 100 else ''}")
            if len(unmapped) > 15:
                print(f"  … and {len(unmapped) - 15} more")

        print(
            f"Rows: seen={stats['rows_seen']}, to import={stats['imported']}, "
            f"empty={stats['empty_rows']}, no answers={stats['no_answers']}, "
            f"skipped duplicate names={stats['skipped_duplicate']}, "
            f"invalid cells={stats['invalid_cells']}"
        )

        wb.close()

        if not commit:
            print("Dry run only — no database changes. Re-run with --commit to insert.")
            return 0

        for rs, resp_list in pending_sessions:
            session.add(rs)
            session.add_all(resp_list)
        await session.commit()
        print(f"Committed {stats['imported']} response sessions.")
        return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Microsoft Forms Excel responses")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_PATH,
        help=f"Path to .xlsx export (default: {DEFAULT_PATH})",
    )
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: active sheet)")
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Write to database (default is dry-run)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview only (default unless --commit)",
    )
    parser.add_argument(
        "--on-duplicate",
        choices=("suffix", "skip"),
        default="suffix",
        help="When respondent name already exists (default: suffix)",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail on invalid option values (default: skip invalid checkbox options leniently)",
    )
    parser.add_argument("--limit", type=int, default=None, help="Import at most N data rows (testing)")
    args = parser.parse_args()

    commit = args.commit and not args.dry_run
    exit_code = asyncio.run(
        run_import(
            args.file,
            sheet=args.sheet,
            commit=commit,
            on_duplicate=args.on_duplicate,
            strict=args.strict,
            limit=args.limit,
        )
    )
    raise SystemExit(exit_code)


if __name__ == "__main__":
    main()
