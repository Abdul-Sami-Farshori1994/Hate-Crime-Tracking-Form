"""Build Excel (.xlsm) exports for admin data portability."""

from __future__ import annotations

import json
import zipfile
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from form_flow import sort_questions
from models import FormPage, Question
from soft_delete import active_form_page, active_question

_XLSX_WORKBOOK_CONTENT_TYPE = (
    b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
_XLSM_WORKBOOK_CONTENT_TYPE = b"application/vnd.ms-excel.sheet.macroEnabled.main+xml"


def _xlsx_bytes_to_xlsm(xlsx_bytes: bytes) -> bytes:
    """openpyxl writes OOXML spreadsheet bytes; Excel requires macro-enabled CT for .xlsm."""
    source = BytesIO(xlsx_bytes)
    destination = BytesIO()
    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(
        destination,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "[Content_Types].xml" and _XLSX_WORKBOOK_CONTENT_TYPE in data:
                data = data.replace(_XLSX_WORKBOOK_CONTENT_TYPE, _XLSM_WORKBOOK_CONTENT_TYPE)
            zout.writestr(info, data)
    return destination.getvalue()


def format_answer_for_cell(answer_value: str | None, question_type: str) -> str:
    if not answer_value:
        return ""
    if question_type == "checkbox":
        try:
            parsed = json.loads(answer_value)
            if isinstance(parsed, list):
                return ", ".join(str(item) for item in parsed)
        except json.JSONDecodeError:
            pass
    return answer_value


def _parse_submitted_at(value: str | None) -> datetime | str:
    if not value:
        return ""
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.replace(tzinfo=None)
        return dt
    except ValueError:
        return value


async def fetch_ordered_questions(db: AsyncSession) -> list[Question]:
    pages_result = await db.execute(
        select(FormPage).where(active_form_page()).order_by(FormPage.order_index, FormPage.id)
    )
    pages = list(pages_result.scalars().all())
    q_result = await db.execute(select(Question).where(active_question()))
    questions = list(q_result.scalars().all())
    return sort_questions(questions, pages)


def build_export_xlsm(
    sessions: list[dict[str, Any]],
    questions: list[Question],
) -> bytes:
    """One row per submission; Microsoft Forms–style meta columns + question columns."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Responses"

    meta_headers = ["Id", "Start time", "Completion time", "Name"]
    headers = meta_headers + [q.question_text for q in questions]
    sheet.append(headers)

    for session in sessions:
        answers_by_qid: dict[int, str] = {}
        for answer in session.get("answers") or []:
            qid = answer.get("question_id")
            if qid is not None:
                answers_by_qid[int(qid)] = format_answer_for_cell(
                    answer.get("answer_value"),
                    str(answer.get("question_type") or ""),
                )

        submitted = _parse_submitted_at(session.get("submitted_at"))
        row: list[Any] = [
            session.get("session_id") or "",
            submitted,
            submitted,
            session.get("respondent_name") or "",
        ]
        for question in questions:
            row.append(answers_by_qid.get(question.id, ""))
        sheet.append(row)

    for index in range(1, len(headers) + 1):
        column = get_column_letter(index)
        sheet.column_dimensions[column].width = min(48, max(12, len(str(headers[index - 1])) + 2))

    buffer = BytesIO()
    workbook.save(buffer)
    return _xlsx_bytes_to_xlsm(buffer.getvalue())


def session_count_from_payload(sessions: list[dict[str, Any]]) -> int:
    return len(sessions)
